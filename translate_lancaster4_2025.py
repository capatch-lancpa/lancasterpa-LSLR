import csv
import datetime
import string
from typing import List, Optional, Union

import openpyxl


# Function to map material codes to material types
def material(material: str) -> Optional[str]:
    """Convert Leadcast material type to DEP Material type

    Args:
        material (str): Leadcast material string

    Returns:
        Optional[str]: DEP material
    """
    """

    """
    old_materials = [
        "A) Lead",  # 0
        "B) Lead-lined galvanized",  # 1
        "C) Galvanized",  # 2
        "D) Copper",  # 3
        "E) Cast iron - lined",  # 4
        "F) Cast iron - unlined",  # 5
        "G) HDPE - high density polyethylene",  # 6
        "H) PVC - polyvinyl chloride",  # 7
        "J) CPVC - chlorine treated PVC",  # 8
        "K) PEX - cross-linked polyethylene",  # 9
        "L) ABS - acrylonitrile butadiene styrene",  # 10
        "M) PB - Polybutylene",  # 11
        "O) Asbestos cement",  # 12
        "P) Other non-lead material",  # 13
        "Q) Unknown - Likely Lead",  # 14
        "R) Unknown - Unlikely Lead",  # 15
        "S) Unknown",  # 16
    ]
    old_material_map = {
        "LD": old_materials[0],
        "CU": old_materials[3],
        "BR": old_materials[13],  # add in comments "Brass"
        "DI": old_materials[13],  # add in comments "Ductile Iron"
        "PVC": old_materials[7],
        "CI": old_materials[5],
        "GALV": old_materials[2],
        "UNK-NL": old_materials[15],
        "UNK": old_materials[16],
        "HDPE": old_materials[6],
        "PE": old_materials[9],
        "PL": old_materials[13],  # add in comments "Plastic"
        "AC": old_materials[12],
    }
    new_materials = [
        "A) LEAD",  # 0
        "B) LEAD-LINED GALVANIZED",  # 1
        "C) GALVANIZED",  # 2
        "D) COPPER",  # 3
        "E) CAST IRON - LINED",  # 4
        "F) CAST IRON - UNLINED",  # 5
        "G) HDPE - HIGH DENSITY POLYETHYLENE",  # 6
        "H) PVC - POLYVINYL CHLORIDE",  # 7
        "I) BRASS",  # 8
        "J) CPVC - CHLORINE TREATED PVC",  # 9
        "K) PEX - CROSS-LINKED POLYETHYLENE",  # 10
        "L) ABS - ACRYLONITRILE BUTADIENE STYRENE",  # 11
        "M) PB - POLYBUTYLENE",  # 12
        "N) DUCTILE IRON",  # 13
        "O) ASBESTOS CEMENT",  # 14
        "P) OTHER NON-LEAD MATERIAL",  # 15
        "S) UNKNOWN",  # 16
        "T) UNKNOWN - NOT LEAD",  # 17
    ]
    new_material_map = {
        "LD": new_materials[0],
        "CU": new_materials[3],
        "BR": new_materials[8],  # add in comments "Brass"
        "DI": new_materials[13],  # add in comments "Ductile Iron"
        "PVC": new_materials[7],
        "CI": new_materials[5],
        "GALV": new_materials[2],
        "UNK-NL": new_materials[17],
        "UNK": new_materials[16],
        "HDPE": new_materials[6],
        "PE": new_materials[17],
        "PL": new_materials[17],  # add in comments "Plastic"
        "AC": new_materials[17],
    }
    return new_material_map.get(material, None)


# Function to map installation dates to predefined ranges
def install_date_range(date: str) -> Optional[str]:
    date_ranges = [
        "A) Pre-1901",
        "B) 1901 - 1910",
        "C) 1911 - 1920",
        "D) 1921 - 1930",
        "E) 1931 - 1940",
        "F) 1941 - 1950",
        "G) 1951 - 1960",
        "H) 1961 - 1970",
        "J) 1971 - 1980",
        "K) 1981 - 1990",
        "L) 1991 - 2000",
        "M) 2001 - 2010",
        "O) 2011 - 2020",
        "P) 2021 - 2030",
    ]

    if not date:
        return None

    # Convert date string to date object
    try:
        utility_install_date = datetime.datetime.strptime(date, "%m/%d/%Y").date()
    except ValueError:
        return None

    # Map date ranges
    date_mapping = [
        (datetime.date(1901, 1, 1), datetime.date(1910, 12, 31), date_ranges[1]),
        (datetime.date(1911, 1, 1), datetime.date(1920, 12, 31), date_ranges[2]),
        (datetime.date(1921, 1, 1), datetime.date(1930, 12, 31), date_ranges[3]),
        (datetime.date(1931, 1, 1), datetime.date(1940, 12, 31), date_ranges[4]),
        (datetime.date(1941, 1, 1), datetime.date(1950, 12, 31), date_ranges[5]),
        (datetime.date(1951, 1, 1), datetime.date(1960, 12, 31), date_ranges[6]),
        (datetime.date(1961, 1, 1), datetime.date(1970, 12, 31), date_ranges[7]),
        (datetime.date(1971, 1, 1), datetime.date(1980, 12, 31), date_ranges[8]),
        (datetime.date(1981, 1, 1), datetime.date(1990, 12, 31), date_ranges[9]),
        (datetime.date(1991, 1, 1), datetime.date(2000, 12, 31), date_ranges[10]),
        (datetime.date(2001, 1, 1), datetime.date(2010, 12, 31), date_ranges[11]),
        (datetime.date(2011, 1, 1), datetime.date(2020, 12, 31), date_ranges[12]),
        (datetime.date(2021, 1, 1), datetime.date(2030, 12, 31), date_ranges[13]),
    ]

    # Determine the date range for the given installation date
    if utility_install_date < datetime.date(1901, 1, 1):
        return date_ranges[0]

    for start_date, end_date, label in date_mapping:
        if start_date <= utility_install_date <= end_date:
            return label

    return None


def increment_label(index):
    """Generate a label (A, B, ..., Z, AA, AB, ..., AZ, BA, ...) for duplicates."""
    label = ""
    while index >= 0:
        label = string.ascii_uppercase[index % 26] + label
        index = index // 26 - 1
    return label


def capitalize_address(street: str) -> str:
    """Capitalize the first letter of each word in the street address."""
    return street.title()


def split_verification_dates(field_from_leadcst, output_additional_comments) -> str:
    """DEP requires that only one date exists in their 'Date of Field Verification' field. This function will split the data coming from leadcast, keep one, and add the remaining to the 'Additional Comments' field

    Args:
        field_from_leadcst (_type_): Either 'Utility Verification date' or 'Private Verification Date'
            Yes, 'Utility Verification date' is correct the date is lowercase. In the future this case might need to be handled if it is made uppercase
        output_additional_comments (_type_): Either one or the other 'Additional Comments' fields in DEP output
    """

    split_dates = field_from_leadcst.split(" | ")

    for i, split_date in enumerate(split_dates):
        if i == 0:
            continue
        else:
            output_additional_comments

    return split_dates[0]


def translate(input_data):
    output_data = []

    # Dictionary to track the count of addresses
    address_count = {}

    # for row in input_data:
    for row in (r for r in input_data if r.get("PWS ID") != "TRAINING"):
        new_row_dict = {
            "UNIQUE SERVICE LINE ID": None,
            "REPLACEMENT DATE": None,
            "SPLIT LINE": None,
            "STREET ADDRESS": None,
            "STREET ADDRESS 2": None,
            "CITY/TOWNSHIP": None,
            "ZIP CODE": None,
            "SCHOOL": None,
            "CHILDCARE": None,
            ###
            "SEGMENT 1 MATERIAL": None,
            "EVER PREVIOUSLY LEAD": None,
            "LEAD CONNECTOR UPSTREAM": None,
            "INSTALLATION DECADE": None,
            "INSTALLATION DATE": None,
            "DIAMETER (IN INCHES)": None,
            "NON-LEAD VERIFICATION 1": None,
            "NON-LEAD VERIFICATION 2": None,
            "FIELD VERIFICATION DATE": None,
            "COMMENTS": None,
            ###
            "SEGMENT 2 MATERIAL": None,
            "LEAD CONNECTOR UPSTREAM_2": None,
            "INSTALLATION DECADE_2": None,
            "INSTALLATION DATE_2": None,
            "NON-LEAD VERIFICATION 1_2": None,
            "NON-LEAD VERIFICATION 2_2": None,
            "FIELD VERIFICATION DATE_2": None,
            "COMMENTS_2": None,
            ###
            "SERVICE LINE CONNECTED TO": None,
            "INORGANIC POE TREATMENT PRESENT": None,
            "INTERIOR PLUMBING": None,
            "LCRI SAMPLING SITE": None,
            ###
            "NUMBER OF CONNECTORS": None,
        }
        ###################################
        ## Service Line Basic Information
        ###################################
        # UNIQUE SERVICE LINE ID
        new_row_dict["UNIQUE SERVICE LINE ID"] = row["ID"]

        # REPLACEMENT DATE
        new_row_dict["REPLACEMENT DATE"] = None

        # SPLIT LINE
        new_row_dict["SPLIT LINE"] = "YES"

        # Street Address 1
        street = row["Street"]
        capitalized_address = capitalize_address(street)
        new_row_dict["STREET ADDRESS"] = capitalize_address(street)

        if capitalized_address in address_count:
            address_count[capitalized_address] += 1
            # Generate the increment label (A, B, AA, etc.) based on the occurrence count
            new_row_dict["STREET ADDRESS 2"] = increment_label(
                address_count[capitalized_address] - 2
            )  # Start from A
        else:
            address_count[capitalized_address] = 1
            new_row_dict["STREET ADDRESS 2"] = (
                None  # First occurrence of this street, no suffix
            )

        # City or Township
        new_row_dict["CITY/TOWNSHIP"] = row["City"]

        # Zip Code
        new_row_dict["ZIP CODE"] = row["Zipcode"]

        # SCHOOL
        var = ["NO", "YES - ELEMENTARY", "YES - SECONDARY", "YES - ALL GRADES"]
        if row["Building Type"] == "Elementary School":
            new_row_dict["SCHOOL"] = var[1]
        elif row["Building Type"] == "School Non-Elementary":
            new_row_dict["SCHOOL"] = var[2]
        else:
            new_row_dict["SCHOOL"] = var[0]

        # Childcare Facility?
        var = ["NO", "YES"]
        if row["Building Type"] in [
            "Day Care",
            "Residential & In-Home Day Care",
        ]:
            new_row_dict["CHILDCARE"] = var[1]
        else:
            new_row_dict["CHILDCARE"] = var[0]

        ###################################
        ## System-Owned Portion of Service Line
        ###################################
        comments_ut = []
        # Material

        # Updated Material hierarchy for selection
        material_priority = ["LD", "GALV", "UNK", "UNK-NL", "CU", "PL"]

        chosen_material = None
        # Material handling (only split if "|" is found)
        if "|" in row["Utility Materials"]:
            system_materials = row["Utility Materials"].split(" | ")
            for priority_material in material_priority:
                if priority_material in system_materials:
                    chosen_material = material(priority_material)
                    break
            if not chosen_material:
                chosen_material = material(
                    "UNK-NL"
                )  # Default to UNK-NL if no match in hierarchy
            new_row_dict["SEGMENT 1 MATERIAL"] = chosen_material
        else:
            new_row_dict["SEGMENT 1 MATERIAL"] = material(
                row["Utility Materials"]
            )  # Treat it as a list with one element if no "|"

        # Was Material Ever Previously Lead?
        if row["Utility Previously Lead"] == "Yes":
            new_row_dict["EVER PREVIOUSLY LEAD"] = "YES"
        elif row["Utility Previously Lead"] == "No":
            new_row_dict["EVER PREVIOUSLY LEAD"] = "NO"
        else:
            new_row_dict["EVER PREVIOUSLY LEAD"] = "NOT SURE"

        # Lead Pigtail, Gooseneck or Connector Upstream?
        if row["Connector Materials"] == "LD":
            new_row_dict["LEAD CONNECTOR UPSTREAM"] = "YES"
        elif row["Connector Materials"] != "LD" and row["Connector Materials"] != "UNK":
            new_row_dict["LEAD CONNECTOR UPSTREAM"] = "NO"
        else:
            new_row_dict["LEAD CONNECTOR UPSTREAM"] = "NOT SURE"

        # Installation Date Range
        utility_installation_date = (
            datetime.datetime.strptime(
                row["Utility Installation Dates"], "%m/%d/%Y"
            ).date()
            if row["Utility Installation Dates"]
            else None
        )
        new_row_dict["INSTALLATION DECADE"] = install_date_range(
            row["Utility Installation Dates"]
        )
        # Installation Date Specific
        new_row_dict["INSTALLATION DATE"] = row["Utility Installation Dates"]

        # "Diameter (in inches)"
        if row["Utility Diameters"] != "99":
            new_row_dict["DIAMETER (IN INCHES)"] = row["Utility Diameters"]

        ################ NEW ################

        """
        # Verification method priority
        
        1. Field verification [Visual Inspection, CCTV, Mechanical Excavation]
        2. Diameter > 2"
        3. Installation Date After Lead Ban
        4. Stats/Modeling [B) MODELING/STATISTICAL ANALYSIS]
        5. Records Review
        
        
        Records - Other
        -Field Inspection
        -Diameter > 2"
        -Installation Date After Lead Ban
        -Records Validation with Field Inspection
        Installation Records
        Other
        -Field Inspection | Installation Date After Lead Ban
        -Records - Other | Field Inspection

        """
        ### Utility Material Method
        utility_verification_methods = []
        if row["Utility Status"] != "Lead Status Unknown":
            if "Field Inspection" in row["Utility Material Method"]:
                utility_verification_methods.append(
                    "E) VISUAL INSPECTION AT 1 ACCESS POINT"
                )
            if 'Diameter > 2"' in row["Utility Material Method"]:
                utility_verification_methods.append(
                    "D) OTHER - ENTER IN COMMENTS FIELD"
                )
                comments_ut.append("Diameter greater than 2 inches")
            if "Installation Date After Lead Ban" in row["Utility Material Method"]:
                utility_verification_methods.append("A) RECORDS REVIEW")
                comments_ut.append("Installation date after lead ban")
            if (
                row["Water Main Install Year"]
                and int(row["Water Main Install Year"]) >= 2012
            ):
                utility_verification_methods.append("O) HIGH CONFIDENCE IN RECORDS")
                comments_ut.append("Water main installed after lead ban")
            if (
                row["Predict Score Utility"]
                and float(row["Predict Score Utility"]) <= 0.1
            ):
                utility_verification_methods.append("B) MODELING/STATISTICAL ANALYSIS")
                comments_ut.append("Predictive model indicates low likelihood of lead")
            if (
                "Records - Other" in row["Utility Material Method"]
                or "Installation Records" in row["Utility Material Method"]
            ):
                utility_verification_methods.append("A) RECORDS REVIEW")
                comments_ut.append("Records review indicates non-lead material")

        if len(utility_verification_methods) >= 1:
            new_row_dict["NON-LEAD VERIFICATION 1"] = utility_verification_methods[0]
        if len(utility_verification_methods) >= 2:
            new_row_dict["NON-LEAD VERIFICATION 2"] = utility_verification_methods[1]

        ################ NEW ################
        # Date of Field Verification
        utility_most_recent_date = None
        if row["Utility Field Verified"] == "Yes" and row["Utility Verification date"]:
            verification_dates = [
                d for d in row["Utility Verification date"].split(" | ")
            ]
            utility_most_recent_date = max(
                verification_dates,
                key=lambda d: datetime.datetime.strptime(d, "%m/%d/%Y"),
            )
            new_row_dict["FIELD VERIFICATION DATE"] = utility_most_recent_date

        # Additional Comments for System-Owned
        if row["Utility Materials"] in ["DI", "BR", "PL"]:
            comments_ut.append(f"Material: {row['Utility Materials']}")

        ################ NEW ################
        # Append Utility Notes if present
        if (
            "Utility side installation date is estimated from installation date of nearest water main"
            in row.get("Utility Notes")
        ):
            comments_ut.append(row["Utility Notes"])

        new_row_dict["COMMENTS"] = " | ".join(comments_ut) if comments_ut else None

        ###################################
        ## Customer-Owned Portion of Service Line
        ###################################
        comments_priv = []
        # Material
        chosen_material = None
        # Material handling (only split if "|" is found)
        if "|" in row["Private Materials"]:
            system_materials = row["Private Materials"].split(" | ")
            for priority_material in material_priority:
                if priority_material in system_materials:
                    chosen_material = material(priority_material)
                    break
            if not chosen_material:
                chosen_material = material(
                    "UNK-NL"
                )  # Default to UNK-NL if no match in hierarchy
            new_row_dict["SEGMENT 2 MATERIAL"] = chosen_material
        else:
            new_row_dict["SEGMENT 2 MATERIAL"] = material(
                row["Private Materials"]
            )  # Treat it as a list with one element if no "|"

        # Lead Pigtail, Gooseneck or Connector Upstream?
        if row["Connector Materials"] == "LD":
            new_row_dict["LEAD CONNECTOR UPSTREAM_2"] = f"YES"
        elif row["Connector Materials"] != "LD" and row["Connector Materials"] != "UNK":
            new_row_dict["LEAD CONNECTOR UPSTREAM_2"] = f"NO"
        else:
            new_row_dict["LEAD CONNECTOR UPSTREAM_2"] = f"NOT SURE"

        # Installation Date Range
        private_installation_date = (
            datetime.datetime.strptime(
                row["Private Installation Dates"], "%m/%d/%Y"
            ).date()
            if row["Private Installation Dates"]
            else None
        )
        new_row_dict["INSTALLATION DECADE_2"] = install_date_range(
            row["Private Installation Dates"]
        )
        # Installation Date Specific
        new_row_dict["INSTALLATION DATE_2"] = row["Private Installation Dates"]

        """
        # Verification method priority
        
        1. Field verification [Visual Inspection, CCTV, Mechanical Excavation]
        2. Diameter > 2"
        3. Installation Date After Lead Ban
        4. Stats/Modeling [B) MODELING/STATISTICAL ANALYSIS]
        5. Records Review
        """
        # Private Material Method
        private_verification_methods = []
        if row["Private Status"] != "Lead Status Unknown":
            if "Field Inspection" in row["Private Material Method"]:
                private_verification_methods.append(
                    "E) VISUAL INSPECTION AT 1 ACCESS POINT"
                )
            if 'Diameter > 2"' in row["Private Material Method"]:
                private_verification_methods.append(
                    "D) OTHER - ENTER IN COMMENTS FIELD"
                )
                comments_priv.append("Diameter greater than 2 inches")
            if "Installation Date After Lead Ban" in row["Private Material Method"]:
                private_verification_methods.append("A) RECORDS REVIEW")
                comments_priv.append("Installation date after lead ban")
            if (
                row["Water Main Install Year"]
                and int(row["Water Main Install Year"]) >= 2012
            ):
                private_verification_methods.append("O) HIGH CONFIDENCE IN RECORDS")
                comments_priv.append("Water main installed after lead ban")
            if (
                row["Predict Score Private"]
                and float(row["Predict Score Private"]) <= 0.1
            ):
                private_verification_methods.append("B) MODELING/STATISTICAL ANALYSIS")
                comments_priv.append(
                    "Predictive model indicates low likelihood of lead"
                )
            if (
                "Records - Other" in row["Private Material Method"]
                or "Installation Records" in row["Private Material Method"]
            ):
                private_verification_methods.append("A) RECORDS REVIEW")
                comments_priv.append("Records review indicates non-lead material")

        if len(private_verification_methods) >= 1:
            new_row_dict["NON-LEAD VERIFICATION 1_2"] = private_verification_methods[0]
        if len(private_verification_methods) >= 2:
            new_row_dict["NON-LEAD VERIFICATION 2_2"] = private_verification_methods[1]

        # Date of Field Verification
        private_most_recent_date = None
        if row["Private Field Verified"] == "Yes" and row["Private Verification Date"]:
            verification_dates = [
                d for d in row["Private Verification Date"].split(" | ")
            ]
            private_most_recent_date = max(
                verification_dates,
                key=lambda d: datetime.datetime.strptime(d, "%m/%d/%Y"),
            )
            new_row_dict["FIELD VERIFICATION DATE_2"] = private_most_recent_date

        # Additional Comments for Customer-Owned
        if row["Private Materials"] in ["DI", "BR", "PL"]:
            comments_priv.append(f"Material: {row['Private Materials']}")

        # Append Private Notes if present
        # if row.get("Private Notes"):
        #     comments_priv.append(row["Private Notes"])

        new_row_dict["COMMENTS_2"] = (
            " | ".join(comments_priv) if comments_priv else None
        )

        ###################################
        ## Information to Assign Tap Monitoring Tiering
        ###################################
        # "SERVICE LINE CONNECTED TO"

        if row["Building Type"] == "Single-Family":
            new_row_dict["SERVICE LINE CONNECTED TO"] = f"S) SINGLE FAMILY RESIDENCE"
        elif row["Building Type"] == "Multi-Family":
            new_row_dict["SERVICE LINE CONNECTED TO"] = f"M) MULTI FAMILY RESIDENCE"
        else:
            new_row_dict["SERVICE LINE CONNECTED TO"] = f"O) BUILDING/OTHER"

        # INORGANIC POE TREATMENT PRESENT
        if row["POE Filter"] == "Unknown":
            new_row_dict["INORGANIC POE TREATMENT PRESENT"] = f"NOT SURE"
        elif row["POE Filter"] == "Yes":
            new_row_dict["INORGANIC POE TREATMENT PRESENT"] = f"YES"
        elif row["POE Filter"] == "No":
            new_row_dict["INORGANIC POE TREATMENT PRESENT"] = f"NO"
        else:
            new_row_dict["INORGANIC POE TREATMENT PRESENT"] = f"NOT SURE"

        # INTERIOR PLUMBING
        if (
            row["Lead Solder Present"] == "Unknown"
            and row["Other Fittings Containing Lead"] == "Unknown"
            and row["Plumbing Material"] == "Unknown"
            and row["Plumbing Contains Lead Solder"] == "Unknown"
        ):
            new_row_dict["INTERIOR PLUMBING"] = f"NOT SURE"
        else:
            if row["Plumbing Material"] == "LD":
                new_row_dict["INTERIOR PLUMBING"] = f"IS LEAD"
            if (
                not new_row_dict["INTERIOR PLUMBING"]
                and row["Plumbing Material"] == "GALV"
            ):
                new_row_dict["INTERIOR PLUMBING"] = f"IS GALVANIZED"
            if (
                not new_row_dict["INTERIOR PLUMBING"]
                and row["Lead Solder Present"] == "Yes"
            ):
                new_row_dict["INTERIOR PLUMBING"] = f"CONTAINS LEAD SOLDER"
            if (
                not new_row_dict["INTERIOR PLUMBING"]
                and row["Lead Solder Present"] == "No"
                and row["Other Fittings Containing Lead"] == "No"
                and row["Plumbing Material"] not in ["LD", "GALV"]
                and row["Plumbing Contains Lead Solder"] == "No"
            ):
                new_row_dict["INTERIOR PLUMBING"] = f"NO LEAD OR GALVANIZED PRESENT"

        # # Current LCR Sampling Site?
        # if row["Sample Site Status"] == "Yes":
        #     new_row_dict["LCRI SAMPLING SITE"] = f"Yes"
        # else:
        #     new_row_dict["LCRI SAMPLING SITE"] = f"No"

        # Store the modified row
        output_data.append(new_row_dict)
    return output_data


def translate_to_csv(input_file, output_file):

    # Open the input CSV file for reading
    with open(input_file, mode="r", encoding="utf-8") as infile:
        reader = csv.DictReader(infile)

        # Open the output CSV file for writing
        with open(output_file, mode="w", newline="", encoding="utf-8") as outfile:
            data = translate(reader)
            # header = []
            header = data[0].keys() if data else []
            writer = csv.DictWriter(outfile, fieldnames=header)
            writer.writeheader()
            for row in data:
                # Write the modified row to the output CSV
                writer.writerow(row)

    print(f"Translation complete. Data saved to {output_file}")


def translate_to_xlsm(input_csv, input_xlsm, output_xlsm):

    # Open the input CSV file for reading
    with open(input_csv, mode="r", encoding="utf-8") as infile:
        reader = csv.DictReader(infile)

        data = translate(reader)

        # Open an existing Excel file or create a new one
        try:
            workbook = openpyxl.load_workbook(input_xlsm, keep_vba=True)
            print(f"File '{input_xlsm}' opened successfully.")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            print(f"File '{input_xlsm}' not found, creating a new one.")

        worksheet = workbook["Detailed Inventory"]
        # E9 starting cell in blank inventory
        # start_row = 9
        # start_col = 5
        curr_row = 10
        curr_col = 5

        for row in data:
            for val in row.values():
                worksheet.cell(row=curr_row, column=curr_col, value=val)
                curr_col += 1
            curr_col = 5
            curr_row += 1

        workbook.save(output_xlsm)


# Example usage
input_csv = "LancasterPA_inventory-export_20251203202621.csv"  # Replace with your input CSV file
output_csv = "translated_output.csv"  # Replace with the output CSV file
translate_to_csv(input_csv, output_csv)

input_xlsm = "SERVICE_LINE_INVENTORY_FORM_2025.xlsm"
output_xlsm = "output_v4.xlsm"
# translate_to_xlsm(input_csv, input_xlsm, output_xlsm)
