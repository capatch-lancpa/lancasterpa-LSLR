import csv
import datetime
import string
from typing import List, Optional, Union

import openpyxl


# Function to map material codes to material types
def material(material: str) -> Optional[str]:
    materials = [
        "A) Lead",  # 0
        "B) Lead-lined galvanized",  # 1
        "C) Galvanized",  # 2
        "D) Copper",  # 3
        "E) Cast iron - lined",  # 4
        "F) Cast iron - unlined",  # 5
        "G) HDPE - high density polyethylene",  # 6
        "H) PVC - polyvinyl chloride",  # 7
        "J) CPVC - chlorine treated PVC",  # 8
        "K) PEX - cross-linked polyethylene",  # 9
        "L) ABS - acrylonitrile butadiene styrene",  # 10
        "M) PB - Polybutylene",  # 11
        "O) Asbestos cement",  # 12
        "P) Other non-lead material",  # 13
        "Q) Unknown - Likely Lead",  # 14
        "R) Unknown - Unlikely Lead",  # 15
        "S) Unknown",  # 16
    ]
    material_map = {
        "LD": materials[0],
        "CU": materials[3],
        "BR": materials[13],  # add in comments "Brass"
        "DI": materials[13],  # add in comments "Ductile Iron"
        "PVC": materials[7],
        "CI": materials[5],
        "GALV": materials[2],
        "UNK-NL": materials[13],
        "UNK": materials[16],
        "HDPE": materials[6],
        "PE": materials[9],
        "PL": materials[13],  # add in comments "Plastic"
        "AC": materials[12],
    }
    return material_map.get(material, None)


# Function to map installation dates to predefined ranges
def install_date_range(date: str) -> Optional[str]:
    date_ranges = [
        "A) Pre-1901",
        "B) 1901 - 1910",
        "C) 1911 - 1920",
        "D) 1921 - 1930",
        "E) 1931 - 1940",
        "F) 1941 - 1950",
        "G) 1951 - 1960",
        "H) 1961 - 1970",
        "J) 1971 - 1980",
        "K) 1981 - 1990",
        "L) 1991 - 2000",
        "M) 2001 - 2010",
        "O) 2011 - 2020",
        "P) 2021 - 2030",
    ]

    if not date:
        return None

    # Convert date string to date object
    try:
        utility_install_date = datetime.datetime.strptime(date, "%m/%d/%Y").date()
    except ValueError:
        return None

    # Map date ranges
    date_mapping = [
        (datetime.date(1901, 1, 1), datetime.date(1910, 12, 31), date_ranges[1]),
        (datetime.date(1911, 1, 1), datetime.date(1920, 12, 31), date_ranges[2]),
        (datetime.date(1921, 1, 1), datetime.date(1930, 12, 31), date_ranges[3]),
        (datetime.date(1931, 1, 1), datetime.date(1940, 12, 31), date_ranges[4]),
        (datetime.date(1941, 1, 1), datetime.date(1950, 12, 31), date_ranges[5]),
        (datetime.date(1951, 1, 1), datetime.date(1960, 12, 31), date_ranges[6]),
        (datetime.date(1961, 1, 1), datetime.date(1970, 12, 31), date_ranges[7]),
        (datetime.date(1971, 1, 1), datetime.date(1980, 12, 31), date_ranges[8]),
        (datetime.date(1981, 1, 1), datetime.date(1990, 12, 31), date_ranges[9]),
        (datetime.date(1991, 1, 1), datetime.date(2000, 12, 31), date_ranges[10]),
        (datetime.date(2001, 1, 1), datetime.date(2010, 12, 31), date_ranges[11]),
        (datetime.date(2011, 1, 1), datetime.date(2020, 12, 31), date_ranges[12]),
        (datetime.date(2021, 1, 1), datetime.date(2030, 12, 31), date_ranges[13]),
    ]

    # Determine the date range for the given installation date
    if utility_install_date < datetime.date(1901, 1, 1):
        return date_ranges[0]

    for start_date, end_date, label in date_mapping:
        if start_date <= utility_install_date <= end_date:
            return label

    return None


# Function to map field methods to predefined options
def field_method(method: str) -> Optional[str]:
    field_methods = [
        "E) Visual inspection at existing access point",
        "F) CCTV inspection inside pipe - full length",
        "G) CCTV inspection outside pipe - at curb box",
        "H) Mechanical excavation - 1 location",
        "J) Mechanical excavation - 2 locations",
        "K) Mechanical excavation - 3+ locations",
        "L) Other - enter in Comments field",
    ]
    if method == "Visual Inspection":
        return field_methods[0]
    return None


def non_field_method(method) -> str:
    var = [
        "A) Records review",
        "B) Modeling/statistical analysis",
        "C) Water sampling (no CCT)",
        "D) Other - enter in Comments field",
    ]
    if method in [
        "Records Validation",
        "Records Invalidation",
        "Installation Date After Lead Ban",
        'Diameter > 2"',
        "Replacement Record",
        "Records - Other",
        "Installation Records",
    ]:
        return var[0]
    elif method in ["Predictive Model", "Statistical Analysis"]:
        return var[1]
    elif method == "Other":
        return var(3)
    else:
        return None


def increment_label(index):
    """Generate a label (A, B, ..., Z, AA, AB, ..., AZ, BA, ...) for duplicates."""
    label = ""
    while index >= 0:
        label = string.ascii_uppercase[index % 26] + label
        index = index // 26 - 1
    return label


def capitalize_address(street: str) -> str:
    """Capitalize the first letter of each word in the street address."""
    return street.title()


def split_verification_dates(field_from_leadcst, output_additional_comments) -> str:
    """DEP requires that only one date exists in their 'Date of Field Verification' field. This function will split the data coming from leadcast, keep one, and add the remaining to the 'Additional Comments' field

    Args:
        field_from_leadcst (_type_): Either 'Utility Verification date' or 'Private Verification Date'
            Yes, 'Utility Verification date' is correct the date is lowercase. In the future this case might need to be handled if it is made uppercase
        output_additional_comments (_type_): Either one or the other 'Additional Comments' fields in DEP output
    """

    split_dates = field_from_leadcst.split(" | ")

    for i, split_date in enumerate(split_dates):
        if i == 0:
            continue
        else:
            output_additional_comments

    return split_dates[0]


def translate(input_data):
    output_data = []

    # Dictionary to track the count of addresses
    address_count = {}

    # for row in input_data:
    for row in (r for r in input_data if r.get("PWS ID") != "TRAINING"):
        new_row_dict = {
            "Unique Service Line ID (Required)": None,
            "Record Type": None,
            "Date Replacement Completed": None,
            "Ownership Type": None,
            "Street Address 1": None,
            "Street Address 2": None,
            "City or Township": None,
            "Zip Code": None,
            "School?": None,
            "Childcare Facility?": None,
            "[Utility] Material": None,
            "[Utility] Was Material Ever Previously Lead?": None,
            "[Utility] Lead Pigtail, Gooseneck or Connector Upstream?": None,
            "[Utility] Installation Date Range": None,
            "[Utility] Installation Date Specific": None,
            "[Utility] Diameter (in inches)": None,
            "[Utility]1 Basis of Material Classification - Non-Field Method": None,
            "[Utility]2 Basis of Material Classification - Non-Field Method": None,
            "[Utility] Basis of Material Classification - Field Method": None,
            "[Utility] Date of Field Verification": None,
            "[Utility] Additional Comments": None,
            "[Private] Material": None,
            "[Private] Lead Pigtail, Gooseneck or Connector Upstream?": None,
            "[Private] Installation Date Range": None,
            "[Private] Installation Date Specific": None,
            "[Private]1 Basis of Material Classification - Non-Field Method": None,
            "[Private]2 Basis of Material Classification - Non-Field Method": None,
            "[Private] Basis of Material Classification - Field Method": None,
            "[Private] Date of Field Verification": None,
            "[Private] Additional Comments": None,
            "Service Line Connected To:": None,
            "POE Treatment Present?": None,
            "Interior Building Plumbing Contains Lead Solder?": None,
            "Current LCR Sampling Site?": None,
        }
        ###################################
        ## Service Line Basic Information
        ###################################
        # Unique Service Line ID (Required)
        new_row_dict["Unique Service Line ID (Required)"] = row["ID"]

        # Record Type
        var = ["Initial", "Update", "Add", "Inactive"]
        new_row_dict["Record Type"] = var[0]

        # Date Replacement Completed
        # Skip

        # Ownership Type
        var = ["Joint", "System", "Customer"]
        new_row_dict["Ownership Type"] = var[0]

        id_value = row["ID"]
        # Street Address 1
        street = row["Street"]
        new_row_dict["Street Address 1"] = capitalize_address(street)

        # Check if the ID has a suffix letter at the end
        if id_value and id_value[-1].isalpha():
            new_row_dict["Street Address 2"] = id_value[
                -1
            ].upper()  # Extract the letter suffix

        # # Street Address 2 (Increment A-Z, AA-ZZ for duplicates)
        else:
            if street in address_count:
                address_count[street] += 1
                # Generate the increment label (A, B, AA, etc.) based on the occurrence count
                new_row_dict["Street Address 2"] = increment_label(
                    address_count[street] - 2
                )  # Start from A
            else:
                address_count[street] = 1
                new_row_dict["Street Address 2"] = (
                    None  # First occurrence of this street, no suffix
                )

        # City or Township
        new_row_dict["City or Township"] = row["City"]

        # Zip Code
        new_row_dict["Zip Code"] = row["Zipcode"]

        # School?
        var = ["No", "Yes - Elementary", "Yes - Secondary", "Yes - All Grades"]
        if row["Building Type"] == "Elementary School":
            new_row_dict["School?"] = var[1]
        elif row["Building Type"] == "School Non-Elementary":
            new_row_dict["School?"] = var[2]
        else:
            new_row_dict["School?"] = var[0]

        # Childcare Facility?
        var = ["No", "Yes"]
        if row["Building Type"] in [
            "Day Care",
            "Residential & In-Home Day Care",
        ]:
            new_row_dict["School?"] = var[1]
        else:
            new_row_dict["School?"] = var[0]

        ###################################
        ## System-Owned Portion of Service Line
        ###################################
        comments_ut = []
        # Material

        # Updated Material hierarchy for selection
        material_priority = ["LD", "GALV", "UNK", "UNK-NL", "CU", "PL"]

        chosen_material = None
        # Material handling (only split if "|" is found)
        if "|" in row["Utility Materials"]:
            system_materials = row["Utility Materials"].split(" | ")
            for priority_material in material_priority:
                if priority_material in system_materials:
                    chosen_material = material(priority_material)
                    break
            if not chosen_material:
                chosen_material = material(
                    "UNK-NL"
                )  # Default to UNK-NL if no match in hierarchy
            new_row_dict["[Utility] Material"] = chosen_material
        else:
            new_row_dict["[Utility] Material"] = material(
                row["Utility Materials"]
            )  # Treat it as a list with one element if no "|"

        # Was Material Ever Previously Lead?
        if row["Utility Previously Lead"] == "Yes":
            new_row_dict["[Utility] Was Material Ever Previously Lead?"] = f"Yes"
        elif row["Utility Previously Lead"] == "No":
            new_row_dict["[Utility] Was Material Ever Previously Lead?"] = f"No"
        elif row["Utility Previously Lead"] == "Unknown":
            new_row_dict["[Utility] Was Material Ever Previously Lead?"] = f"Not sure"
        else:
            new_row_dict["[Utility] Was Material Ever Previously Lead?"] = None

        # Lead Pigtail, Gooseneck or Connector Upstream?
        if row["Connector Materials"] == "LD":
            new_row_dict["[Utility] Lead Pigtail, Gooseneck or Connector Upstream?"] = (
                f"Yes"
            )
        elif row["Connector Materials"] != "LD" and row["Connector Materials"] != "UNK":
            new_row_dict["[Utility] Lead Pigtail, Gooseneck or Connector Upstream?"] = (
                f"No"
            )
        else:
            new_row_dict["[Utility] Lead Pigtail, Gooseneck or Connector Upstream?"] = (
                f"Not sure"
            )

        # Installation Date Handling (only split if "|" is found)
        if "|" in row["Utility Installation Dates"]:
            utility_dates = [
                d
                for d in row["Utility Installation Dates"].split(" | ")
                if d != "1/1/1970"
            ]
            most_recent_date = max(
                utility_dates, key=lambda d: datetime.datetime.strptime(d, "%m/%d/%Y")
            )
            # Installation Date Range
            new_row_dict["[Utility] Installation Date Range"] = install_date_range(
                most_recent_date
            )  # Use most recent date for range
            # Installation Date Specific
            new_row_dict["[Utility] Installation Date Specific"] = (
                most_recent_date  # Most recent date specific
            )
        else:
            # Installation Date Range
            new_row_dict["[Utility] Installation Date Range"] = install_date_range(
                row["Utility Installation Dates"]
            )
            # Installation Date Specific
            new_row_dict["[Utility] Installation Date Specific"] = row[
                "Utility Installation Dates"
            ]

        # "Diameter (in inches)"
        if row["Utility Diameters"] != "99":
            new_row_dict["[Utility] Diameter (in inches)"] = row["Utility Diameters"]

        # "Basis of Material Classification - Non-Field Method"
        if row["Utility Material Method"] == "Installation Date After Lead Ban":
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Utility]1 Basis of Material Classification - Non-Field Method"
            ] = f"A) Records Review"
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Utility]2 Basis of Material Classification - Non-Field Method"
            ] = f"D) Other - enter in Comments field"
            comments_ut.append(
                f"We have high confidence in this record that the service line is non-lead due to the City of Reading lead ban in 1976."
            )
        elif row["Utility Material Method"] == 'Diameter > 2"':
            new_row_dict[
                "[Utility]1 Basis of Material Classification - Non-Field Method"
            ] = f"A) Records Review"
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Utility]2 Basis of Material Classification - Non-Field Method"
            ] = f"D) Other - enter in Comments field"
            comments_ut.append(
                f"We have high confidence in this record from the internal records that the service line diameter is > 2 inches."
            )
        elif row["Utility Material Method"] == "Records - Other":
            new_row_dict[
                "[Utility]1 Basis of Material Classification - Non-Field Method"
            ] = f"A) Records Review"
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Utility]2 Basis of Material Classification - Non-Field Method"
            ] = None
            comments_ut.append(
                f"We have high confidence in this record from the internal records."
            )
        else:
            new_row_dict[
                "[Utility]1 Basis of Material Classification - Non-Field Method"
            ] = non_field_method(row["Utility Verification Method"])
            new_row_dict[
                "[Utility]2 Basis of Material Classification - Non-Field Method"
            ] = None

        # "Basis of Material Classification - Field Method"
        if row["Utility Field Verified"] == "Yes":
            new_row_dict[
                "[Utility] Basis of Material Classification - Field Method"
            ] = field_method(row["Utility Verification Method"])

        # Date of Field Verification
        if row["Utility Field Verified"] == "Yes":
            verification_dates = [
                d for d in row["Utility Verification date"].split(" | ")
            ]
            most_recent_date = max(
                verification_dates,
                key=lambda d: datetime.datetime.strptime(d, "%m/%d/%Y"),
            )
            new_row_dict["[Utility] Date of Field Verification"] = most_recent_date

        # Additional Comments for System-Owned
        if row["Utility Materials"] in ["DI", "BR", "PL"]:
            comments_ut.append(f"Material: {row['Utility Materials']}")

        # Append Utility Notes if present
        if row.get("Utility Notes"):
            comments_ut.append(row["Utility Notes"])

        new_row_dict["[Utility] Additional Comments"] = (
            " | ".join(comments_ut) if comments_ut else None
        )

        ###################################
        ## Customer-Owned Portion of Service Line
        ###################################
        comments_priv = []
        # Material
        chosen_material = None
        # Material handling (only split if "|" is found)
        if "|" in row["Private Materials"]:
            system_materials = row["Private Materials"].split(" | ")
            for priority_material in material_priority:
                if priority_material in system_materials:
                    chosen_material = material(priority_material)
                    break
            if not chosen_material:
                chosen_material = material(
                    "UNK-NL"
                )  # Default to UNK-NL if no match in hierarchy
            new_row_dict["[Private] Material"] = chosen_material
        else:
            new_row_dict["[Private] Material"] = material(
                row["Private Materials"]
            )  # Treat it as a list with one element if no "|"

        # Lead Pigtail, Gooseneck or Connector Upstream?
        if row["Connector Materials"] == "LD":
            new_row_dict["[Private] Lead Pigtail, Gooseneck or Connector Upstream?"] = (
                f"Yes"
            )
        elif row["Connector Materials"] != "LD" and row["Connector Materials"] != "UNK":
            new_row_dict["[Private] Lead Pigtail, Gooseneck or Connector Upstream?"] = (
                f"No"
            )
        else:
            new_row_dict["[Private] Lead Pigtail, Gooseneck or Connector Upstream?"] = (
                f"Not sure"
            )

        # Installation Date Handling (only split if "|" is found)
        if "|" in row["Private Installation Dates"]:
            private_dates = [
                d
                for d in row["Private Installation Dates"].split(" | ")
                if d != "1/1/1970"
            ]
            most_recent_date = max(
                private_dates, key=lambda d: datetime.datetime.strptime(d, "%m/%d/%Y")
            )
            # Installation Date Range
            new_row_dict["[Private] Installation Date Range"] = install_date_range(
                most_recent_date
            )  # Use most recent date for range
            # Installation Date Specific
            new_row_dict["[Private] Installation Date Specific"] = (
                most_recent_date  # Most recent date specific
            )
        else:
            # Installation Date Range
            new_row_dict["[Private] Installation Date Range"] = install_date_range(
                row["Private Installation Dates"]
            )
            # Installation Date Specific
            new_row_dict["[Private] Installation Date Specific"] = row[
                "Private Installation Dates"
            ]

        # "Basis of Material Classification - Non-Field Method"
        if row["Private Material Method"] == "Installation Date After Lead Ban":
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Private]1 Basis of Material Classification - Non-Field Method"
            ] = f"A) Records Review"
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Private]2 Basis of Material Classification - Non-Field Method"
            ] = f"D) Other - enter in Comments field"
            comments_priv.append(
                f"We have high confidence in this record that the service line is non-lead due to the City of Reading lead ban in 1976."
            )
        elif row["Private Material Method"] == 'Diameter > 2"':
            new_row_dict[
                "[Private]1 Basis of Material Classification - Non-Field Method"
            ] = f"A) Records Review"
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Private]2 Basis of Material Classification - Non-Field Method"
            ] = f"D) Other - enter in Comments field"
            comments_priv.append(
                f"We have high confidence in this record from the internal records that the service line diameter is > 2 inches."
            )
        elif row["Private Material Method"] == "Records - Other":
            new_row_dict[
                "[Private]1 Basis of Material Classification - Non-Field Method"
            ] = f"A) Records Review"
            # "Basis of Material Classification - Non-Field Method"
            new_row_dict[
                "[Private]2 Basis of Material Classification - Non-Field Method"
            ] = None
            comments_priv.append(
                f"We have high confidence in this record from the internal records."
            )
        else:
            new_row_dict[
                "[Private]1 Basis of Material Classification - Non-Field Method"
            ] = non_field_method(row["Private Verification Method"])
            new_row_dict[
                "[Private]2 Basis of Material Classification - Non-Field Method"
            ] = None

        # "Basis of Material Classification - Field Method"
        new_row_dict["[Private] Basis of Material Classification - Field Method"] = (
            field_method(row["Private Verification Method"])
        )

        # Date of Field Verification
        if row["Private Field Verified"] == "Yes":
            verification_dates = [
                d for d in row["Private Verification Date"].split(" | ")
            ]
            most_recent_date = max(
                verification_dates,
                key=lambda d: datetime.datetime.strptime(d, "%m/%d/%Y"),
            )
            new_row_dict["[Private] Date of Field Verification"] = most_recent_date

        # Additional Comments for Customer-Owned
        if row["Private Materials"] in ["DI", "BR", "PL"]:
            comments_priv.append(f"Material: {row['Private Materials']}")

        # Append Private Notes if present
        if row.get("Private Notes"):
            comments_priv.append(row["Private Notes"])

        new_row_dict["[Private] Additional Comments"] = (
            " | ".join(comments_priv) if comments_priv else None
        )

        ###################################
        ## Information to Assign Tap Monitoring Tiering
        ###################################
        # "Service Line Connected To:"
        """
        var = [
            "S) Single family residence",
            "M) Multi family residence",
            "O) Building/Other",
        ]
        """
        if row["Building Type"] == "Single-Family":
            new_row_dict["Service Line Connected To:"] = f"S) Single family residence"
        elif row["Building Type"] == "Multi-Family":
            new_row_dict["Service Line Connected To:"] = f"M) Multi family residence"
        else:
            new_row_dict["Service Line Connected To:"] = f"O) Building/Other"

        # POE Treatment Present?
        if row["POE Filter"] == "Unknown":
            new_row_dict["POE Treatment Present?"] = f"Not sure"
        elif row["POE Filter"] == "Yes":
            new_row_dict["POE Treatment Present?"] = f"Yes"
        elif row["POE Filter"] == "No":
            new_row_dict["POE Treatment Present?"] = f"No"
        else:
            new_row_dict["POE Treatment Present?"] = f"Not sure"

        # Interior Building Plumbing Contains Lead Solder?
        if row["Plumbing Contains Lead Solder"] == "Unknown":
            new_row_dict["Interior Building Plumbing Contains Lead Solder?"] = (
                f"Not sure"
            )
        elif row["Plumbing Contains Lead Solder"] == "Yes":
            new_row_dict["Interior Building Plumbing Contains Lead Solder?"] = f"Yes"
        elif row["Plumbing Contains Lead Solder"] == "No":
            new_row_dict["Interior Building Plumbing Contains Lead Solder?"] = f"No"
        else:
            new_row_dict["Interior Building Plumbing Contains Lead Solder?"] = (
                f"Not sure"
            )

        # Current LCR Sampling Site?
        if row["Sample Site Status"] == "Yes":
            new_row_dict["Current LCR Sampling Site?"] = f"Yes"
        else:
            new_row_dict["Current LCR Sampling Site?"] = f"No"

        # Store the modified row
        output_data.append(new_row_dict)
    return output_data


def translate_to_csv(input_file, output_file):

    # Open the input CSV file for reading
    with open(input_file, mode="r", encoding="utf-8") as infile:
        reader = csv.DictReader(infile)

        # Open the output CSV file for writing
        with open(output_file, mode="w", newline="", encoding="utf-8") as outfile:
            data = translate(reader)
            header = [
                "Unique Service Line ID (Required)",
                "Record Type",
                "Date Replacement Completed",
                "Ownership Type",
                "Street Address 1",
                "Street Address 2",
                "City or Township",
                "Zip Code",
                "School?",
                "Childcare Facility?",
                "[Utility] Material",
                "[Utility] Was Material Ever Previously Lead?",
                "[Utility] Lead Pigtail, Gooseneck or Connector Upstream?",
                "[Utility] Installation Date Range",
                "[Utility] Installation Date Specific",
                "[Utility] Diameter (in inches)",
                "[Utility]1 Basis of Material Classification - Non-Field Method",
                "[Utility]2 Basis of Material Classification - Non-Field Method",
                "[Utility] Basis of Material Classification - Field Method",
                "[Utility] Date of Field Verification",
                "[Utility] Additional Comments",
                "[Private] Material",
                "[Private] Lead Pigtail, Gooseneck or Connector Upstream?",
                "[Private] Installation Date Range",
                "[Private] Installation Date Specific",
                "[Private]1 Basis of Material Classification - Non-Field Method",
                "[Private]2 Basis of Material Classification - Non-Field Method",
                "[Private] Basis of Material Classification - Field Method",
                "[Private] Date of Field Verification",
                "[Private] Additional Comments",
                "Service Line Connected To:",
                "POE Treatment Present?",
                "Interior Building Plumbing Contains Lead Solder?",
                "Current LCR Sampling Site?",
            ]
            writer = csv.DictWriter(outfile, fieldnames=header)
            writer.writeheader()
            for row in data:
                # Write the modified row to the output CSV
                writer.writerow(row)

    print(f"Translation complete. Data saved to {output_file}")


def translate_to_xlsm(input_csv, input_xlsm, output_xlsm):

    # Open the input CSV file for reading
    with open(input_csv, mode="r", encoding="utf-8") as infile:
        reader = csv.DictReader(infile)

        data = translate(reader)

        # Open an existing Excel file or create a new one
        try:
            workbook = openpyxl.load_workbook(input_xlsm, keep_vba=True)
            print(f"File '{input_xlsm}' opened successfully.")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            print(f"File '{input_xlsm}' not found, creating a new one.")

        worksheet = workbook["Detailed Inventory"]
        # E9 starting cell in blank inventory
        # start_row = 9
        # start_col = 5
        curr_row = 10
        curr_col = 5

        for row in data:
            for val in row.values():
                worksheet.cell(row=curr_row, column=curr_col, value=val)
                curr_col += 1
            curr_col = 5
            curr_row += 1

        workbook.save(output_xlsm)


# Example usage
input_csv = (
    "Inventory-LancasterPA-1728501219510.csv"  # Replace with your input CSV file
)
output_csv = "translated_output.csv"  # Replace with the output CSV file
# translate_to_csv(input_csv, output_csv)

input_xlsm = "SERVICE_LINE_INVENTORY_FORM.xlsm"
output_xlsm = "output.xlsm"
translate_to_xlsm(input_csv, input_xlsm, output_xlsm)
