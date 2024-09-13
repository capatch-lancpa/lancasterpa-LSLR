import csv
import datetime

import openpyxl


def material(material) -> str:
    var = [
        "A) Lead",  # 0
        "B) Lead-lined galvanized",  # 1
        "C) Galvanized",  # 2
        "D) Copper",  # 3
        "E) Cast iron - lined",  # 4
        "F) Cast iron - unlined",  # 5
        "G) HDPE - high density polyethylene",  # 6
        "H) PVC - polyvinyl chloride",  # 7
        "J) CPVC - chlorine treated PVC",  # 8
        "K) PEX - cross-linked polyethylene",  # 9
        "L) ABS - acrylonitrile butadiene styrene",  # 10
        "M) PB - Polybutylene",  # 11
        "O) Asbestos cement",  # 12
        "P) Other non-lead material",  # 13
        "Q) Unknown - Likely Lead",  # 14
        "R) Unknown - Unlikely Lead",  # 15
        "S) Unknown",  # 16
    ]
    if material == "LD":
        return var[0]
    elif material == "CU":
        return var[3]
    elif material == "BR":
        return var[13]
    elif material == "DI":
        return var[13]
    elif material == "PL":
        return var[7]
    elif material == "CI":
        return var[5]
    elif material == "GALV":
        return var[2]
    elif material == "UNK-NL":
        return var[15]
    elif material == "UNK":
        return var[16]
    else:
        return None


def install_date_range(date) -> str:
    var = [
        "A) Pre-1901",
        "B) 1901 - 1910",
        "C) 1911 - 1920",
        "D) 1921 - 1930",
        "E) 1931 - 1940",
        "F) 1941 - 1950",
        "G) 1951 - 1960",
        "H) 1961 - 1970",
        "J) 1971 - 1980",
        "K) 1981 - 1990",
        "L) 1991 - 2000",
        "M) 2001 - 2010",
        "O) 2011 - 2020",
        "P) 2021 - 2030",
    ]
    if not date:
        return None
    utility_install_date = datetime.datetime.strptime(date, "%m/%d/%Y").date()
    if utility_install_date < datetime.date(1901, 1, 1):
        return var[0]
    elif (
        datetime.date(1901, 1, 1) <= utility_install_date <= datetime.date(1910, 12, 31)
    ):
        return var[1]
    elif (
        datetime.date(1911, 1, 1) <= utility_install_date <= datetime.date(1920, 12, 31)
    ):
        return var[2]
    elif (
        datetime.date(1921, 1, 1) <= utility_install_date <= datetime.date(1930, 12, 31)
    ):
        return var[3]
    elif (
        datetime.date(1931, 1, 1) <= utility_install_date <= datetime.date(1940, 12, 31)
    ):
        return var[4]
    elif (
        datetime.date(1941, 1, 1) <= utility_install_date <= datetime.date(1950, 12, 31)
    ):
        return var[5]
    elif (
        datetime.date(1951, 1, 1) <= utility_install_date <= datetime.date(1960, 12, 31)
    ):
        return var[6]
    elif (
        datetime.date(1961, 1, 1) <= utility_install_date <= datetime.date(1970, 12, 31)
    ):
        return var[7]
    elif (
        datetime.date(1971, 1, 1) <= utility_install_date <= datetime.date(1980, 12, 31)
    ):
        return var[8]
    elif (
        datetime.date(1981, 1, 1) <= utility_install_date <= datetime.date(1990, 12, 31)
    ):
        return var[9]
    elif (
        datetime.date(1991, 1, 1) <= utility_install_date <= datetime.date(2000, 12, 31)
    ):
        return var[10]
    elif (
        datetime.date(2001, 1, 1) <= utility_install_date <= datetime.date(2010, 12, 31)
    ):
        return var[11]
    elif (
        datetime.date(2011, 1, 1) <= utility_install_date <= datetime.date(2020, 12, 31)
    ):
        return var[12]
    elif (
        datetime.date(2021, 1, 1) <= utility_install_date <= datetime.date(2030, 12, 31)
    ):
        return var[13]
    else:
        return None


def field_method(method) -> str:
    var = [
        "E) Visual inspection at existing access point",
        "F) CCTV inspection inside pipe - full length",
        "G) CCTV inspection outside pipe - at curb box",
        "H) Mechanical excavation - 1 location",
        "J) Mechanical excavation - 2 locations",
        "K) Mechanical excavation - 3+ locations",
        "L) Other - enter in Comments field",
    ]
    if method == "Visual Inspection":
        return var[0]
    else:
        return None


def non_field_method(method) -> str:
    var = [
        "A) Records review",
        "B) Modeling/statistical analysis",
        "C) Water sampling (no CCT)",
        "D) Other - enter in Comments field",
    ]
    if method in [
        "Records Validation",
        "Records Invalidation",
        "Installation Date After Lead Ban",
        'Diameter > 2"',
        "Replacement Record",
        "Records - Other",
        "Installation Records",
    ]:
        return var[0]
    elif method in ["Predictive Model", "Statistical Analysis"]:
        return var[1]
    elif method == "Other":
        return var(3)
    else:
        return None


def translate(input_data):
    output_data = []
    for row in input_data:

        new_row = []
        ###################################
        ## Service Line Basic Information
        ###################################
        # Unique Service Line ID (Required)
        new_row.append(row["ID"])

        # Record Type
        var = ["Initial", "Update", "Add", "Inactive"]
        new_row.append(var[0])

        # Date Replacement Completed
        new_row.append(None)

        # Ownership Type
        var = ["Joint", "System", "Customer"]
        new_row.append(var[0])

        # Street Address 1
        new_row.append(row["Street"])

        # Street Address 2
        new_row.append(None)

        # City or Township
        new_row.append(row["City"])

        # Zip Code
        new_row.append(row["Zipcode"])

        # School?
        var = ["No", "Yes - Elementary", "Yes - Secondary", "Yes - All Grades"]
        if row["Building Type"] == "Elementary School":
            new_row.append(var[1])
        elif row["Building Type"] == "School Non-Elementary":
            new_row.append(var[2])
        else:
            new_row.append(var[0])

        # Childcare Facility?
        var = ["No", "Yes"]
        if row["Building Type"] in [
            "Day Care",
            "Residential & In-Home Day Care",
        ]:
            new_row.append(var[1])
        else:
            new_row.append(var[0])

        ###################################
        ## System-Owned Portion of Service Line
        ###################################
        # Material
        new_row.append(material(row["Utility Materials"]))

        # Was Material Ever Previously Lead?
        var = ["Yes", "No", "Not sure"]
        if row["Utility Previously Lead"] == "Yes":
            new_row.append(var[0])
        elif row["Utility Previously Lead"] == "No":
            new_row.append(var[1])
        elif row["Utility Previously Lead"] == "Unknown":
            new_row.append(var[2])
        else:
            new_row.append(None)

        # Lead Pigtail, Gooseneck or Connector Upstream?
        var = ["Yes", "No", "Not sure"]
        new_row.append(None)

        # Installation Date Range
        new_row.append(install_date_range(row["Utility Installation Dates"]))

        # Installation Date Specific
        new_row.append(row["Utility Installation Dates"])

        # "Diameter (in inches)"
        if row["Utility Diameters"] != "99":
            new_row.append(row["Utility Diameters"])
        else:
            new_row.append(None)

        # "Basis of Material Classification - Non-Field Method"
        new_row.append(non_field_method(row["Utility Verification Method"]))

        ##### DUPLICATE?!?!
        # "Basis of Material Classification - Non-Field Method"
        # new_row.append(non_field_method(row["Utility Verification Method"]))
        new_row.append(None)

        # "Basis of Material Classification - Field Method"
        if row["Utility Field Verified"] == "Yes":
            new_row.append(field_method(row["Utility Verification Method"]))
        else:
            new_row.append(None)

        # Date of Field Verification
        if row["Utility Field Verified"] == "Yes":
            new_row.append(row["Utility Verification date"])
        else:
            new_row.append(None)

        # Additional Comments
        new_row.append(None)

        ###################################
        ## Customer-Owned Portion of Service Line
        ###################################
        # Material
        new_row.append(material(row["Private Materials"]))

        # Lead Pigtail, Gooseneck or Connector Upstream?
        var = ["Yes", "No", "Not sure"]
        new_row.append(None)

        # Installation Date Range
        new_row.append(install_date_range(row["Private Installation Dates"]))

        # Installation Date Specific
        new_row.append(row["Private Installation Dates"])

        # "Basis of Material Classification - Non-Field Method"
        new_row.append(non_field_method(row["Private Verification Method"]))

        ##### DUPLICATE?!?!
        # "Basis of Material Classification - Non-Field Method"
        # new_row.append(non_field_method(row["Private Verification Method"]))
        new_row.append(None)

        # "Basis of Material Classification - Field Method"
        new_row.append(field_method(row["Private Verification Method"]))

        # Date of Field Verification
        if row["Private Field Verified"] == "Yes":
            new_row.append(row["Private Verification Date"])
        else:
            new_row.append(None)

        # Additional Comments
        new_row.append(None)

        ###################################
        ## Information to Assign Tap Monitoring Tiering
        ###################################
        # "Service Line Connected To:"
        var = [
            "S) Single family residence",
            "M) Multi family residence",
            "O) Building/Other",
        ]
        if row["Building Type"] == "Single-Family":
            new_row.append(var[0])
        elif row["Building Type"] == "Multi-Family":
            new_row.append(var[1])
        else:
            new_row.append(var[2])

        # POE Treatment Present?
        var = ["Yes", "No", "Not sure"]
        new_row.append(None)

        # Interior Building Plumbing Contains Lead Solder?
        var = ["Yes", "No", "Not sure"]
        new_row.append(None)

        # Current LCR Sampling Site?
        var = ["No", "Yes"]
        new_row.append(None)

        # Check to make sure we have all 34 values
        if len(new_row) != 34:
            break

        # Store the modified row
        output_data.append(new_row)
    return output_data


def translate_to_csv(input_file, output_file):

    # Open the input CSV file for reading
    with open(input_file, mode="r", encoding="utf-8") as infile:
        reader = csv.DictReader(infile)

        # Open the output CSV file for writing
        with open(output_file, mode="w", newline="", encoding="utf-8") as outfile:
            writer = csv.writer(outfile)

            data = translate(reader)
            for row in data:
                # Write the modified row to the output CSV
                writer.writerow(row)

    print(f"Translation complete. Data saved to {output_file}")


def translate_to_xlsm(input_csv, input_xlsm, output_xlsm):

    # Open the input CSV file for reading
    with open(input_csv, mode="r", encoding="utf-8") as infile:
        reader = csv.DictReader(infile)

        data = translate(reader)

        # Open an existing Excel file or create a new one
        try:
            workbook = openpyxl.load_workbook(input_xlsm, keep_vba=True)
            print(f"File '{input_xlsm}' opened successfully.")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            print(f"File '{input_xlsm}' not found, creating a new one.")

        worksheet = workbook["Detailed Inventory"]
        # E9 starting cell in blank inventory
        # start_row = 9
        # start_col = 5
        curr_row = 9
        curr_col = 5

        for row in data:
            for val in row:
                worksheet.cell(row=curr_row, column=curr_col, value=val)
                curr_col += 1
            curr_col = 5
            curr_row += 1

        workbook.save(output_xlsm)


# Example usage
input_csv = "input.csv"  # Replace with your input CSV file
output_csv = "translated_output.csv"  # Replace with the output CSV file
# translate_csv(input_csv, output_csv)

input_xlsm = "SERVICE_LINE_INVENTORY_FORM.xlsm"
output_xlsm = "output.xlsm"
translate_to_xlsm(input_csv, input_xlsm, output_xlsm)
